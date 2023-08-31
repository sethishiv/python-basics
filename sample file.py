'''
price = 1000000
has_good_credit = False

if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"down payment: ${down_payment}") #format of string
'''
#import converter

#comparision operators
'''
name = "Shivani"
 #print(len(name)) to get the no of character ex: 1
if len(name) < 3:                 #condtion 1
    print("name must be at least 3 characters")
elif len(name) > 50:             #condition2  #semicolon must be at the end of elif condition
    print("name must be a maximum of 50 characters.")
else:
    print("name looks good!")    #none of teh above condition then this is displayed 
'''
'''
weight = int(input("weight: "))   #to ask the users (user information) ==>> use input function
unit = input("(L)bs or (k)g: ") #user input
if unit.upper() == "L":
    converted = weight * 0.45 #store it in a variable called converted
    print(f"you are {converted} kilos") #formatted string  and dynamically insert value of converted
else:
    converted = weight/ 0.45  #double slash for floating number
    print(f"you are {converted} pounds") #formatted string
'''
#WHILE LOOPS
'''
i = 1
while i <= 5:
    print('*' * i) 
    i = i + 1  #for 1,2,3,4, loop runs when it is 5 (5+1=6) the condition is false so loop will be terminated
print("done")
'''
#GUESS GAME
'''
secret_number = 6
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
    guess = int(input("guess: "))
    guess_count += 1   # variable is incremented by 1 using the statement which ensures the termination of the loop 
    #If you remove or do not include the guess_count += 1 statement, the value of guess_count will always remain 0, 
    #and the condition guess_count < guess_limit will always be true. As a result, the loop will run infinitely
    if guess == secret_number:
        print("you won")
        break   #this loop will be executed till completion until it exceeds guess limit
else:
    print("sorry, you failed ") #when the above condition becomes false this line is executed
'''
#CAR GAME
"""
command = ""
started = False
quit_flag = False
while not quit_flag:
    command = input("> ").lower() #To ensure that every input line is in lowercase and prints in lowercase, use the lower() method on the user's input command to convert it to lowercase before processing or displaying it
    if command == "start":
        if started:
            print("the car is already started")
            continue #to check whether the car started
        else:
            started = True
        print("the car started ....")
    elif command == "stop":
        if not started:
            print("the car is already stopped.")
            continue #By using the continue statement, the program jumps back to the start of the loop without executing the remaining lines of code within that iteration

        else:
            started =  False      #if the car isnt stopped we need to stop it so use bool false
        print("the car stopped.")
    elif  command == "help":
        print("""
#start- to start the car
#stop- to stop the car
#quit- to quit
""")
    if command == "quit":
        break   # remember to break the loop
        print("sorry, i dont understand that....")
"""
#FOR LOOP
# The process of going through each repetition of the loop is called iteration.
#for loop, you can iterate over a sequence of elements and perform a set of actions on each element.In a while loop, you can continue iterating as long as a certain condition remains true.
"""for item in "Python":
        print(item)"""
#range function
'''for item in range(5,10, 3): #here 3 is the gap between two numbers
        print(item)'''
''' prices = [10,20,30] #calculate the total cost using for loop
total = 0
for price in prices:
        total += price
print(f"total: {total}")
'''
#NESTED LOOP
"""for x in range(4): #outer loop 
   for y in range(3): #when the inner loop completes 0,1,2,
        print(f'({x}, {y})')
"""
""""
numbers = [5,2,5,2,2]
for x_count in numbers:
        output = ''
        for count in range(x_count): #it reads the count of x in the first iteration i,e (5) likewise 2,5,2,2 runs on loop
            output += 'x'
        print(output)
"""
#LIST (mutables, square brackets)
'''
names = ['john','bob','boom','mosh']
names[0] = 'jon'
print(names[0:1])
'''
#find the largest number in a list
''' numbers = [210,200,400,500,100]
max = numbers  [0]
for number in numbers:
        if number > max:
           max = number
print(max)'''
#2D LIST
'''matrix =[
        [1,2,3],
        [4,5,6],
        [7,8,9]
]
print(matrix[0][1])
'''
#sets
'''numbers =  [2,2,4,6,3,4,6,1]
uniques =[]
for number in numbers:
        if number not in uniques:          #The not in operator checks for the absence of the number in the uniques list.
            uniques.append(number)
print(uniques)
'''
#tuples (immutable),parenthesis
'''number = (1,2,3)
number[0] = 10 #you cannot change the value
print(number [0])
'''
#UNPACKING
'''coordinates = (1,2,3)
x = coordinates[0]
y= coordinates [1]
z = coordinates[2]
#replace it with one powerful code
x, y, z = coordinates #same result
print(y)
'''
#dictionaries (key value pairs)
#KEY :  VALUE
'''Name: Shivani
Email: Shivani@gmail.com
Phone: 1234757989'''
'''
customer = {
    "name": "john Smith",
    "age": 30,#add coma to add another key
    "is_verified": True
} #if the key doesnt exist we can supply a default value
customer["name"] = "jack william"     #easy to add new keys
print(customer["name"])
print(customer.get("birthdate", "jan 1 1980"))
'''
#need to print the numbers into number names
''' Phone = input("Phone: " ) #a dictionary allows us to map the key to the value
digits_mapping = {
     "1": "one",
     "2": "two",
     "3": "three",
     "4": "four"
 }
output = ""
for ch in Phone:
     output += digits_mapping.get(ch, "!")  +  " "
     #if you dont have a key set it to default
print(output) '''
#emoji convertor (The code essentially takes a user input, splits it into words, and replaces specific words with corresponding emojis using a dictionary.
'''message = input(">")
words = message.split(" ") #each item in the sentence is a string
emojis = {          #emojis = {...}: This line defines a dictionary called emojis. The dictionary maps certain words (keys) to their corresponding emojis (values). In this example, ":)" maps to "😊" and ":(" maps to "😢"
    ":)": "😊", # type Windows logo key + . (period)
    ":(": "😢"
}
output = ""     #output = "": This line initializes an empty string called output. This variable will be used to store the converted message with emojis.
for word in words: #chk the key- word has value orelse return back the sentence 
    output += emojis.get(word,word) + " " #The line output += emojis.get(word, word) + " " retrieves the corresponding emoji from the emojis dictionary or appends the original word, and then adds it to the output string along with a space.
# even if you dont have the value of the key it can bring it and set it as default after comma you give the key value or default one
print(output)'''

#functions
'''def tell_joke():
    print(' Two fish are in a tank. One says, ‘How do you drive this thing?')
    print('Maybe if we start telling people their brain is an app, they’ll want to use it')

print("start")
tell_joke()
print('finish')'''
#functions and parameters
#keyword arguments gives one word description of the argument
'''def greet_user(first_name, last_name):
    print(f'hi {first_name} {last_name}!')
    print('welcome aboard')'''

'''print("start")
greet_user("john", "Smith")
#greet_user(first_name = "john", "smith") ❌❌
#positional argument should come first after that key argument
greet_user("john", last_name = "Smith") #use keyword arguments to increase the readibility
print("finish")'''

#return statement
'''def square(number):  #square function squares the number
    return number * number'''

#result = square(3)
#print(result)
#instead write this one word code
'''print(square(3))'''
#none is the function which represents the absence of something
'''def square(number):
    print(number * number) # i replaced return instead of print by default python returns statement "none"
print(square(3))'''

'''def emojis_convertor(message):
    words = message.split(" ")  # each item in the sentence is a string
    emojis = {
        # emojis = {...}: This line defines a dictionary called emojis. The dictionary maps certain words (keys) to their corresponding emojis (values). In this example, ":)" maps to "😊" and ":(" maps to "😢"
        ":)": "😊",  # type Windows logo key + . (period)
        ":(": "😢"
    }
    output = ""  # output = "": This line initializes an empty string called output. This variable will be used to store the converted message with emojis.
    for word in words:  # chk the key- word has value orelse return back the sentence
        output += emojis.get(word,word) + " "
    return output

#we need to leave two line space be4 printing the output
message = input(">")
print(emojis_convertor(message))'''
#exception

#when the return statement is encountered immediately the function terminates and control of the program goes back to the place from where it is called

#create a program to add and multiply two numbers


#exception error which crashes the program
'''try:
    age = int(input('age: '))
    print(age)
except ValueError: #doesnt let the program crash
    print('Invalid value')'''


'''try:
    age = int(input("age: "))
    income = 200000
    risk = income / age
    print(age)
    print(risk)
except ZeroDivisionError:
    print("age cannot be 0.")
except ValueError:
    print('Invalid value')'''

 #CLASSES: needs to be capatilised called as pascal language
'''class Point:
    def move(self):
        print("move")
    def draw(self):
        print("draw")'''

#to create an object
'''point1 = Point()
point1.x = 10
point1.y = 20
print(point1.y)
point1.draw() '''

#constructors (is a function that gets called at the time of creating an object)
'''class Point:
    def __init__(self,x,y):  #is used to add new objects
        self.x = x
        self.y = y
    def move(self):
        print("move")
    def draw(self):
        print("draw")

point = Point(10,20)
point.x = 11 #you can easily change the value
print(point.x)'''

#person name talk()
class Person:
    def __init__(self,name):
        self.name = name
    def talk(self):
        print(f"hi, i am {self.name}")
        #print("talk") if you change this code to the above code then you require to write print(john.name)
john = Person("john smith") #to add the name we use the constructor
#print(john.name) because of the change in line 308 ❌❌🚫🚫
john.talk()

bob = Person("Bob Smith")
bob.talk()

Shivani = Person("Shivani sethi")
Shivani.talk()

'''class Dog:
    def walk(self):
        print(walk)
#we need to leave 2 lines space between two classes

class Cat:
    def walk(self):
        print("walk")
#now using INHERITANCE
class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self): #definig a method
        print("bark")
    #pass #python doesnt like empty function so you can write pass to leave it empty


class Cat(Mammal):
    def be_annoying(self): #defining method
        print("annoying")

cat1 = Cat()
cat1.be_annoying '''
#so both dog and cat classes are inheriting both the classes mehod defined in their parent class

#MODULES each file is called module and it should contain all the related functions and classes
#then we can import a module into another module
# in this case, we are importing module
'''import converter
from converter import kg_to_lbs
print(converter.lbs_to_kg(70))'''


#shortcut to change the function name shift f5
#to unhastag a statement ctrl hastag

'''from utils import find_max #instead of importing the whole module we can import only a specidic function from the module

numbers = [10,3,2,4]
maximum = find_max(numbers)
print(maximum)'''

#PACKAGES
'''from ecommerce.shipping import calc_shipping

calc_shipping() #this command is used to import a particular function from the module like shipping
#to import the whole module in order to access all the functions and classes
from ecommerce import shipping #now shipping is like object
#Now the shipping module,
shipping.calc_shipping()'''

#Generating random values
'''import random  #its a build in module in python
for i in range(3):
    print(random.randint(10,20))

memebers = ['shiavni', 'surbhi', 'mom','papa']
leader = random.choice(memebers)
print(leader)'''
#import random

'''class dice():
    def roll(self):
        first = random.randint(1,6)
        second = random.randint(1,6)
        return first, second


dice = dice()
print(dice.roll())'''

#files and directories
#from pathlib import Path
#absolute path - c:\Program
#relative path- we start from the current directory

'''path = Path("emails")
print(path.mkdir()) #to make new dictionary
#boolean answer
path = Path() #to get names of all the .py files present in sysytem
for files  in path.glob('*.py'):
    print(files)'''

'''path = Path() #to get names of all the .py files present in sysytem
for files in path.glob('*'): #for names of directories
    print(files)'''

#Excel spreadsheet
#copywrite sales #ghoga


#make it in transaction file excel download it
'''import openpyxl as xl
xl.load_workbook("transactions.xlsx")
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)
print(cell.value)

for row in range(1,sheet.max_row + 1): #to include the second range number
    print(row)''' #to get cell in the 3rd column
'''
import openpyxl
workbook = openpyxl.load_workbook('')


import os

# Specify the file name and directory path
file_name = 'transactions.xlsx'
directory_path = 'C:/Documents/New folder/'

# Create the file path by joining the directory path and file name
file_path = os.path.join(directory_path, transactions.xlsx)

# Check if the file exists
if os.path.isfile():
    print("The Excel file exists in the specified directory.")
    # You can proceed to load the file and automate it using openpyxl
else:
    print("The Excel file does not exist in the specified directory.")
'''
'''import os
folder_name = 'new folder'
file_name = 'transactions.xlsx'

documents_path = os.path.expanduser('~/Documents')
directory_path = os.path.join(documents_path, folder_name)

# Create the file path by joining the directory path and file name
file_path = os.path.join(directory_path, file_name)

# Check if the file exists'''


"""from openpyxl import Workbook, load_workbook

wb = load_workbook('transactions.xlsx')
ws = wb.active
print(ws)"""


#MACHINE LEARNING in action
'''1. import the data csp file
2. clean the data
3. plit the data into training/test sets
4. create a model 
5.train the model 
6. make the predictions 
7. evaluate and improve '''


# LIBRARIES AND TOOLS
'''1.NUMPY - array
2.PANDAS - data anyalsis data frame
3.Mathplotlib -  2D graphs and plots
4.scikit-learn - decision trees and neural networks '''

#ERROR HANLING :
'''try:
    a = int(input("Enter your number: "))
    print(a + 3)
except Exception as e: #this tells the type of error

    print("some error occurred", e)'''

#FILE I/O : to add string to any file
#context manager:

  

with open('text.txt', 'r') as f:
    s = f.read()
    print(s)
#it runs the text in other file (refractor = delete)'''

 #APPEND TO A FILE repeats the text line

with open('text.txt ', 'a') as f:
     f.write('shiavni is serious\n')

#OOPS - OBJECT ORIENTED PROGRAM
'''classes - is a templete to make object
ex: reservation ka form and object is the barah hua reservation ka form'''

class Employee:
    def __init__(self, name, salary):
      self.name = name
      self.salary = salary


    def getsalary(self):
        return self.salary

rohan = Employee('rohan','3737')
print(rohan.salary)
print(rohan.name)
#to create new objects
shivani  = Employee('shivani' ,'6700000000000')
print(shivani .salary)
print(shivani.name)

#to add more employees we use constructor which automatically runswhen the object is created
#--inint--(self):









































































































































































